from django.shortcuts import render, get_object_or_404
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate
from datetime import timedelta

from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import FacturaEnc,FacturaDet
from fe.models import Empresa


def imprimir_factura_recibo(request,id):
    template_name="fac/factura_one.html"

    enc = get_object_or_404(FacturaEnc, id=id)
    det = FacturaDet.objects.filter(factura=id)
    empresa = Empresa.objects.first()

    context={
        'request':request,
        'enc':enc,
        'detalle':det,
        'empresa':empresa,
    }

    return render(request,template_name,context)


def _contexto_reporte_facturas(f1, f2):
    f1_parsed = parse_date(f1)
    f2_parsed = parse_date(f2)
    f2_con_margen = f2_parsed + timedelta(days=1)

    enc = FacturaEnc.objects.filter(fecha__gte=f1_parsed, fecha__lt=f2_con_margen).order_by('id')

    empresa = Empresa.objects.first()

    return {
        'f1': f1_parsed,
        'f2': f2_parsed,
        'enc': enc,
        'empresa': empresa,
        'fecha_emision': timezone.localtime(timezone.now()),
    }


def imprimir_factura_list(request,f1,f2):
    template_name="fac/facturas_print_all.html"

    context = _contexto_reporte_facturas(f1, f2)
    context['request'] = request
    context['es_pdf'] = False
    context['url_descargar_pdf'] = f"/fac/facturas/imprimir-todas-pdf/{f1}/{f2}"

    return render(request,template_name,context)


def imprimir_factura_list_pdf(request, f1, f2):
    template_name = "fac/facturas_print_all.html"

    context = _contexto_reporte_facturas(f1, f2)
    context['request'] = request
    context['es_pdf'] = True
    context['url_ver_en_pantalla'] = f"/fac/facturas/imprimir-todas/{f1}/{f2}"

    html = render_to_string(template_name, context)

    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = f"reporte_facturas_{f1}_a_{f2}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    resultado = pisa.CreatePDF(html, dest=response)
    if resultado.err:
        return HttpResponse(
            "Ocurrió un error al generar el PDF. Contacte al administrador.",
            status=500
        )

    return response


def imprimir_factura_list_excel(request, f1, f2):
    context = _contexto_reporte_facturas(f1, f2)
    facturas = context['enc']

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    encabezados = ["No.", "Fecha", "Cliente", "Total", "Estado", "Anulada"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for f in facturas:
        estado = f.get_estado_sin_display() if hasattr(f, 'get_estado_sin_display') else f.estado_sin
        ws.append([
            f.id,
            f.fecha.strftime("%d/%m/%Y") if f.fecha else "",
            str(f.cliente),
            f.total,
            estado,
            "Sí" if f.anulado else "No",
        ])
        ws.cell(row=ws.max_row, column=1).number_format = '@'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="facturas_{f1}_a_{f2}.xlsx"'
    wb.save(response)
    return response


def _contexto_reporte_cierre_ventas(f1, f2):
    f1_parsed = parse_date(f1)
    f2_parsed = parse_date(f2)
    f2_con_margen = f2_parsed + timedelta(days=1)

    facturas = FacturaEnc.objects.filter(fecha__gte=f1_parsed, fecha__lt=f2_con_margen)

    por_dia = (
        facturas.annotate(dia=TruncDate('fecha'))
        .values('dia')
        .annotate(
            cantidad=Count('id'),
            cantidad_anuladas=Count('id', filter=Q(anulado=True)),
            monto_activo=Sum('total', filter=Q(anulado=False)),
            monto_anulado=Sum('total', filter=Q(anulado=True)),
        )
        .order_by('dia')
    )

    from .models import CierreDia
    cierres = {
        c.fecha: c
        for c in CierreDia.objects.filter(fecha__gte=f1_parsed, fecha__lte=f2_parsed)
    }

    dias = []
    total_activo = 0
    total_anulado = 0
    total_facturas = 0
    cierres_con_observacion = []

    for row in por_dia:
        fecha = row['dia']
        cierre = cierres.get(fecha)
        monto_activo = round(row['monto_activo'] or 0, 2)
        monto_anulado = round(row['monto_anulado'] or 0, 2)

        dias.append({
            'fecha': fecha,
            'cantidad': row['cantidad'],
            'cantidad_anuladas': row['cantidad_anuladas'],
            'monto_activo': monto_activo,
            'monto_anulado': monto_anulado,
            'cierre': cierre,
        })

        total_activo += monto_activo
        total_anulado += monto_anulado
        total_facturas += row['cantidad']

        if cierre and cierre.estado == CierreDia.ESTADO_CERRADO_CON_PENDIENTES:
            cierres_con_observacion.append(cierre)

    empresa = Empresa.objects.first()

    return {
        'f1': f1_parsed,
        'f2': f2_parsed,
        'dias': dias,
        'total_activo': round(total_activo, 2),
        'total_anulado': round(total_anulado, 2),
        'total_neto': round(total_activo, 2),
        'total_facturas': total_facturas,
        'cierres_con_observacion': cierres_con_observacion,
        'empresa': empresa,
        'fecha_emision': timezone.localtime(timezone.now()),
    }


def reporte_cierre_ventas(request, f1, f2):
    template_name = "fac/cierre_ventas_reporte.html"

    context = _contexto_reporte_cierre_ventas(f1, f2)
    context['request'] = request
    context['es_pdf'] = False
    context['url_descargar_pdf'] = f"/fac/reportes/cierre-ventas/pdf/{f1}/{f2}"

    return render(request, template_name, context)


def reporte_cierre_ventas_pdf(request, f1, f2):
    template_name = "fac/cierre_ventas_reporte.html"

    context = _contexto_reporte_cierre_ventas(f1, f2)
    context['request'] = request
    context['es_pdf'] = True

    html = render_to_string(template_name, context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cierre_ventas_{f1}_a_{f2}.pdf"'
    resultado = pisa.CreatePDF(html, dest=response)
    if resultado.err:
        return HttpResponse("Ocurrió un error al generar el PDF.", status=500)
    return response

def _contexto_cierre_caja(f1, f2):
    """
    Cierre de Caja: desglose de ingresos por forma de pago (Resumen) y
    listado factura por factura AGRUPADO por forma de pago, con
    subtotal por grupo (Detallado). Solo se consideran facturas
    ACTIVAS (no anuladas) -- una factura anulada nunca represento un
    ingreso real de caja.
    """
    from itertools import groupby

    f1_parsed = parse_date(f1)
    f2_parsed = parse_date(f2)
    f2_con_margen = f2_parsed + timedelta(days=1)

    facturas_activas = FacturaEnc.objects.filter(
        fecha__gte=f1_parsed, fecha__lt=f2_con_margen, anulado=False
    )

    resumen_qs = (
        facturas_activas.values('forma_pago')
        .annotate(cantidad=Count('id'), total=Sum('total'))
        .order_by('forma_pago')
    )
    etiquetas = dict(FacturaEnc.FORMA_PAGO_CHOICES)
    resumen = [
        {
            'forma_pago': etiquetas.get(row['forma_pago'], row['forma_pago']),
            'cantidad': row['cantidad'],
            'total': round(row['total'] or 0, 2),
        }
        for row in resumen_qs
    ]
    total_general = round(sum(r['total'] for r in resumen), 2)
    cantidad_general = sum(r['cantidad'] for r in resumen)

    # Detalle agrupado por forma de pago, con subtotal por grupo --
    # requiere ordenar por forma_pago primero para que groupby agrupe
    # correctamente (itertools.groupby solo agrupa elementos
    # consecutivos, no re-ordena por si solo).
    detalle_qs = facturas_activas.select_related('cliente').order_by('forma_pago', 'id')
    detalle_agrupado = []
    for codigo_forma_pago, grupo in groupby(detalle_qs, key=lambda f: f.forma_pago):
        facturas_grupo = list(grupo)
        detalle_agrupado.append({
            'forma_pago': etiquetas.get(codigo_forma_pago, codigo_forma_pago),
            'facturas': facturas_grupo,
            'cantidad': len(facturas_grupo),
            'subtotal': round(sum(f.total for f in facturas_grupo), 2),
        })

    empresa = Empresa.objects.first()

    return {
        'f1': f1_parsed,
        'f2': f2_parsed,
        'resumen': resumen,
        'total_general': total_general,
        'cantidad_general': cantidad_general,
        'detalle_agrupado': detalle_agrupado,
        'empresa': empresa,
        'fecha_emision': timezone.localtime(timezone.now()),
    }

def cierre_caja_resumen(request, f1, f2):
    context = _contexto_cierre_caja(f1, f2)
    context['request'] = request
    context['es_pdf'] = False
    context['url_descargar_pdf'] = f"/fac/reportes/cierre-caja/resumen/pdf/{f1}/{f2}/"
    return render(request, 'fac/cierre_caja_resumen.html', context)

def cierre_caja_resumen_pdf(request, f1, f2):
    context = _contexto_cierre_caja(f1, f2)
    context['request'] = request
    context['es_pdf'] = True
    html = render_to_string('fac/cierre_caja_resumen.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cierre_caja_resumen_{f1}_a_{f2}.pdf"'
    resultado = pisa.CreatePDF(html, dest=response)
    if resultado.err:
        return HttpResponse("Ocurrió un error al generar el PDF.", status=500)
    return response

def cierre_caja_detallado(request, f1, f2):
    context = _contexto_cierre_caja(f1, f2)
    context['request'] = request
    context['es_pdf'] = False
    context['url_descargar_pdf'] = f"/fac/reportes/cierre-caja/detallado/pdf/{f1}/{f2}/"
    return render(request, 'fac/cierre_caja_detallado.html', context)

def cierre_caja_detallado_pdf(request, f1, f2):
    context = _contexto_cierre_caja(f1, f2)
    context['request'] = request
    context['es_pdf'] = True
    html = render_to_string('fac/cierre_caja_detallado.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cierre_caja_detallado_{f1}_a_{f2}.pdf"'
    resultado = pisa.CreatePDF(html, dest=response)
    if resultado.err:
        return HttpResponse("Ocurrió un error al generar el PDF.", status=500)
    return response